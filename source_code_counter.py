#!/usr/bin/env python3

#
# source_code_counter.py
#
# Date    : 2024-04-24
# Auther  : Hirotoshi FUJIBE
# History :
#
# Copyright (c) 2024 Hirotoshi FUJIBE
#

"""
Usage:

    Python.exe source_code_counter.py

Options:

    -h
    --help
        Print this message and exit.
"""

#
# Import Libraries
#
import os
import sys
import getopt
import shutil
import datetime
import openpyxl
from typing import Union
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

#
# Input, Output
#
IN_DIR = '.\\input'
OUT_DIR = '.\\output'
# IN_SRC_ROOT = 'D:\\Developments\\PyCharmProjects\\source_code_counter\\input'  # noqa
IN_SRC_ROOT = '.\\input'
IN_SRC_RELATIVE = '\\src'
IN_EXCEL = IN_DIR + '\\source_code_counter_list_template.xlsx'
OUT_EXCEL = OUT_DIR + '\\source_code_counter_list.xlsx'
OUT_DEBUG = OUT_DIR + '\\debug.txt'
OUT_SHEET = 'Source Code Counter List'
ENCODINGS = ['utf-8', 'shift-jis', 'gb2312']
IGNORE_EXTENDS = ['.dat', '.ini']

#
# Excel Cell Position (1 Origin)
#
CELL_ROW_OFFSET = 4
CELL_COL_NO = 2
CELL_COL_PATH = 3
CELL_COL_FILE = 4
CELL_COL_EXT = 5
CELL_COL_LINES = 6
CELL_COL_STEPS = 7

#
# Constants
#

# For Excel
ALIGN_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_LEFT_NO_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=False)
ALIGN_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='top', wrap_text=True)
FONT_MEIRYO = Font(name='Meiryo UI', size=10, color='000000')
FONT_MEIRYO_GRAY = Font(name='Meiryo UI', size=10, color='C0C0C0')
FONT_MEIRYO_BOLD = Font(name='Meiryo UI', size=10, color='000000', bold=True)
FILL_BRIGHT_GRAY = PatternFill(patternType='solid', fgColor='EBECF0')
NUMBER_FORMAT = '#,##0_ '
BORDER_ALL = Border(
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'))

# For Message
MSG_ERROR = 'error'
MSG_NORMAL = 'normal'

# For Characters
# 'noqa' means ignore PEP 8 warning.
SIGN_EXCLAMATION        = '!'   # !     # noqa: E221
SIGN_DOUBLE_QUOTATION   = '"'   # '     # noqa: E221
SIGN_HASH               = '#'   # #     # noqa: E221
SIGN_DOLLARS            = '$'   # $     # noqa: E221
SIGN_PERCENT            = '%'   # %     # noqa: E221
SIGN_AMPERSAND          = '&'   # &     # noqa: E221
SIGN_SINGLE_QUOTATION   = '\''  # '     # noqa: E221
SIGN_LEFT_PAREN         = '('   # (     # noqa: E221
SIGN_RIGHT_PAREN        = ')'   # )     # noqa: E221
SIGN_MINUS              = '-'   # -     # noqa: E221
SIGN_EQUAL              = '='   # =     # noqa: E221
SIGN_CARET              = '^'   # ^     # noqa: E221
SIGN_TILDE              = '~'   # ~     # noqa: E221
SIGN_BACK_SLASH         = '\\'  # \     # noqa: E221
SIGN_VERTICAL           = '|'   # |     # noqa: E221
SIGN_AT                 = '@'   # @     # noqa: E221
SIGN_BACK_APOSTROPHE    = '`'   # `     # noqa: E221
SIGN_LEFT_BRACKET       = '['   # [     # noqa: E221
SIGN_LEFT_BRACE         = '{'   # {     # noqa: E221
SIGN_SEMI_COLON         = ';'   # ;     # noqa: E221
SIGN_PLUS               = '+'   # +     # noqa: E221
SIGN_COLON              = ':'   # :     # noqa: E221
SIGN_ASTERISK           = '*'   # *     # noqa: E221
SIGN_RIGHT_BRACKET      = ']'   # ]     # noqa: E221
SIGN_RIGHT_BRACE        = '}'   # }     # noqa: E221
SIGN_COMMA              = ','   # ,     # noqa: E221
SIGN_LESS_THAN          = '<'   # <     # noqa: E221
SIGN_PERIOD             = '.'   # .     # noqa: E221
SIGN_GREATER_THAN       = '>'   # >     # noqa: E221
SIGN_SLASH              = '/'   # /     # noqa: E221
SIGN_QUESTION           = '?'   # ?     # noqa: E221
SIGN_UNDERSCORE         = '_'   # _     # noqa: E221

CH_SIGNS_OF_PYTHON = [
    SIGN_EXCLAMATION,
    SIGN_DOLLARS,
    SIGN_PERCENT,
    SIGN_AMPERSAND,
    SIGN_LEFT_PAREN,
    SIGN_RIGHT_PAREN,
    SIGN_EQUAL,
    SIGN_CARET,
    SIGN_TILDE,
    SIGN_VERTICAL,
    SIGN_AT,
    SIGN_BACK_APOSTROPHE,
    SIGN_LEFT_BRACKET,
    SIGN_LEFT_BRACE,
    SIGN_SEMI_COLON,
    SIGN_PLUS,
    SIGN_COLON,
    SIGN_RIGHT_BRACKET,
    SIGN_RIGHT_BRACE,
    SIGN_COMMA,
    SIGN_LESS_THAN,
    SIGN_PERIOD,
    SIGN_GREATER_THAN,
    SIGN_QUESTION,
    SIGN_MINUS,
    SIGN_SLASH,
    SIGN_ASTERISK,
]

CH_SIGNS_OF_JAVA = [
    SIGN_EXCLAMATION,
    SIGN_DOLLARS,
    SIGN_PERCENT,
    SIGN_AMPERSAND,
    SIGN_LEFT_PAREN,
    SIGN_RIGHT_PAREN,
    SIGN_EQUAL,
    SIGN_CARET,
    SIGN_TILDE,
    SIGN_VERTICAL,
    SIGN_AT,
    SIGN_BACK_APOSTROPHE,
    SIGN_LEFT_BRACKET,
    SIGN_LEFT_BRACE,
    SIGN_SEMI_COLON,
    SIGN_PLUS,
    SIGN_COLON,
    SIGN_RIGHT_BRACKET,
    SIGN_RIGHT_BRACE,
    SIGN_COMMA,
    SIGN_LESS_THAN,
    SIGN_PERIOD,
    SIGN_GREATER_THAN,
    SIGN_QUESTION,
    SIGN_MINUS,
    SIGN_HASH,
]

CH_SIGNS_OF_SQL = [
    SIGN_EXCLAMATION,
    SIGN_DOLLARS,
    SIGN_PERCENT,
    SIGN_AMPERSAND,
    SIGN_LEFT_PAREN,
    SIGN_RIGHT_PAREN,
    SIGN_EQUAL,
    SIGN_CARET,
    SIGN_TILDE,
    SIGN_VERTICAL,
    SIGN_AT,
    SIGN_BACK_APOSTROPHE,
    SIGN_LEFT_BRACKET,
    SIGN_LEFT_BRACE,
    SIGN_SEMI_COLON,
    SIGN_PLUS,
    SIGN_COLON,
    SIGN_RIGHT_BRACKET,
    SIGN_RIGHT_BRACE,
    SIGN_COMMA,
    SIGN_LESS_THAN,
    SIGN_PERIOD,
    SIGN_GREATER_THAN,
    SIGN_QUESTION,
    SIGN_HASH,
]

CH_DELIMITERS = [' ', '\t', '　']


# Write Excel
class WriteExcel:

    def __init__(self, in_excel: str, out_excel: str, out_sheet: str) -> None:
        shutil.copy(in_excel, out_excel)
        self._wb = openpyxl.load_workbook(out_excel)
        self._sheet = self._wb[out_sheet]
        self._row_offset = CELL_ROW_OFFSET
        self._row = 0
        self._out_excel = out_excel
        return

    def next_row(self) -> None:
        self._row += 1
        return

    def get_row(self) -> int:
        return self._row

    def write_cell(self, i_col: int, i_value: Union[int, str],
                   i_align: int = None, i_font: Font = None, i_format: str = None) -> None:
        self._sheet.cell(row=self._row_offset + self._row, column=i_col).border = BORDER_ALL
        if i_value is not None:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).value = i_value
        if i_align is not None:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).alignment = i_align
        if i_font is not None:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).font = i_font
            if i_font == FONT_MEIRYO_BOLD:
                self._sheet.cell(row=self._row_offset + self._row, column=i_col).fill = FILL_BRIGHT_GRAY
        else:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).font = FONT_MEIRYO
        if i_format is not None:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).number_format = i_format
        return

    def close(self) -> None:
        self._wb.save(self._out_excel)
        self._wb.close()
        return


# Scan Python File
def scan_python_file(full_path_file: str, fp) -> (int, int, str):

    if fp is not None:
        fp.write('%s\n' % full_path_file)

    for enc in ENCODINGS:

        num_lines = 0
        num_steps = 0
        num_double_q = 0
        is_comment = False

        file = open(full_path_file, 'r', encoding=enc)

        while True:

            try:
                str_line = file.readline()
            except Exception:       # noqa
                file.close()
                break

            # End of Data
            if not str_line:
                file.close()
                return num_lines, num_steps, MSG_NORMAL

            num_lines += 1

            str_line = str_line.rstrip('\n')  # for Display
            str_comp = str_line.strip()

            pos_current = 0
            pos_end = len(str_comp)

            tokens = []
            str_token = ''
            str_const = ''
            is_single_q = False
            is_escape = False
            is_ope = False

            while pos_current < pos_end:

                ch = str_comp[pos_current]

                # Inside of Comment
                if is_comment:
                    if ch == SIGN_DOUBLE_QUOTATION:
                        num_double_q += 1
                        if num_double_q == 3:
                            num_double_q = 0
                            is_comment = False

                # Inside of String Constant
                elif str_const != '':
                    if is_escape:
                        str_const += ch
                        is_escape = False
                    elif ch == SIGN_BACK_SLASH:
                        str_const += SIGN_BACK_SLASH
                        is_escape = True
                    else:
                        if is_single_q:
                            if ch == SIGN_SINGLE_QUOTATION:
                                str_const += SIGN_SINGLE_QUOTATION
                                tokens.append(str_const)
                                str_const = ''
                                is_ope = True
                                is_single_q = False
                            else:
                                str_const += ch
                        elif num_double_q == 1:
                            if ch == SIGN_DOUBLE_QUOTATION:
                                str_const += SIGN_DOUBLE_QUOTATION
                                tokens.append(str_const)
                                str_const = ''
                                is_ope = True
                                num_double_q = 0
                            else:
                                str_const += ch

                # '\''
                elif ch == SIGN_SINGLE_QUOTATION:
                    if num_double_q == 1:
                        str_const = SIGN_DOUBLE_QUOTATION + SIGN_SINGLE_QUOTATION
                    elif num_double_q == 2:
                        str_const = SIGN_DOUBLE_QUOTATION + SIGN_DOUBLE_QUOTATION
                        tokens.append(str_const)
                        str_const = SIGN_SINGLE_QUOTATION
                        is_ope = True
                        num_double_q = 0
                        is_single_q = True
                        is_escape = False
                    else:
                        str_const = SIGN_SINGLE_QUOTATION
                        is_single_q = True
                        is_escape = False

                # '"'
                elif ch == SIGN_DOUBLE_QUOTATION:
                    num_double_q += 1
                    if num_double_q == 3:
                        is_comment = True
                        num_double_q = 0
                    is_escape = False

                # '#'
                elif ch == SIGN_HASH:
                    if num_double_q == 1:
                        str_const = SIGN_DOUBLE_QUOTATION + SIGN_HASH
                        continue
                    elif num_double_q == 2:
                        str_const = SIGN_DOUBLE_QUOTATION + SIGN_DOUBLE_QUOTATION
                        tokens.append(str_const)
                        str_const = ''
                        num_double_q = 0
                    else:
                        if str_token != '':
                            tokens.append(str_token)
                            str_token = ''
                            is_ope = True
                            num_double_q = 0
                    break

                # '\'
                elif ch == SIGN_BACK_SLASH:
                    if num_double_q == 1:
                        str_const = SIGN_DOUBLE_QUOTATION
                        continue
                    elif num_double_q == 2:
                        str_const = SIGN_DOUBLE_QUOTATION + SIGN_DOUBLE_QUOTATION
                        tokens.append(str_const)
                        str_const = ''
                        is_ope = True
                        num_double_q = 0
                    elif num_double_q == 3:
                        is_comment = True
                        num_double_q = 0
                    else:
                        if str_token != '':
                            tokens.append(str_token)
                            str_token = ''
                            is_ope = True
                            num_double_q = 0
                    break

                # ' ', '\t', '　'
                elif ch in CH_DELIMITERS:
                    if num_double_q == 1:
                        str_const = SIGN_DOUBLE_QUOTATION + ch
                    elif num_double_q == 2:
                        str_const = SIGN_DOUBLE_QUOTATION + SIGN_DOUBLE_QUOTATION
                        tokens.append(str_const)
                        str_const = ''
                        is_ope = True
                        num_double_q = 0
                    else:
                        if str_token != '':
                            tokens.append(str_token)
                            str_token = ''
                            is_ope = True
                            num_double_q = 0

                # Sign Marks
                elif ch in CH_SIGNS_OF_PYTHON:
                    if num_double_q == 1:
                        str_const = SIGN_DOUBLE_QUOTATION + ch
                    elif num_double_q == 2:
                        str_const = SIGN_DOUBLE_QUOTATION + SIGN_DOUBLE_QUOTATION
                        tokens.append(str_const)
                        str_const = ''
                        is_ope = True
                        num_double_q = 0
                        tokens.append(ch)
                    else:
                        if str_token != '':
                            tokens.append(str_token)
                            str_token = ''
                        tokens.append(ch)
                        is_ope = True

                # Letters, Numbers
                else:
                    if num_double_q == 1:
                        str_const = SIGN_DOUBLE_QUOTATION + ch
                    else:
                        str_token += ch

                pos_current += 1

            # End of One Line
            if str_const != '':
                tokens.append(str_const)
                is_ope = True
            elif str_token != '':
                tokens.append(str_token)
                is_ope = True

            if fp is not None:
                strs = ''
                spc = ''
                for cnt, val in enumerate(tokens):
                    strs += spc + '[' + val + ']'
                    spc = ' '
                fp.write('%s %5d: %s\n' % ('|' if is_ope else ' ', num_lines, strs))

            if is_ope:
                num_steps += 1

        # End of All lines

    print('file encoding error in %s' % full_path_file, file=sys.stderr)
    return 0, 0, MSG_ERROR


# Scan Java File
def scan_java_file(full_path_file: str, fp) -> (int, int, str):

    if fp is not None:
        fp.write('%s\n' % full_path_file)

    for enc in ENCODINGS:

        num_lines = 0
        num_steps = 0
        is_comment = False

        file = open(full_path_file, 'r', encoding=enc)

        while True:

            try:
                str_line = file.readline()
            except Exception:       # noqa
                file.close()
                break

            # End of Data
            if not str_line:
                file.close()
                return num_lines, num_steps, MSG_NORMAL

            num_lines += 1

            str_line = str_line.rstrip('\n')  # for Display
            str_comp = str_line.strip()

            pos_current = 0
            pos_end = len(str_comp)

            tokens = []
            str_token = ''
            str_const = ''
            is_double_q = False
            is_single_q = False
            is_escape = False
            is_slash = False
            is_asterisk = False
            is_ope = False

            while pos_current < pos_end:

                ch = str_comp[pos_current]

                # Inside of Comment
                if is_comment:
                    # before '*' was appeared
                    if is_asterisk:
                        # '*/'
                        if ch == SIGN_SLASH:
                            is_comment = False
                        # '**'
                        elif ch == SIGN_ASTERISK:
                            is_asterisk = True
                        # '*?"
                        else:
                            is_asterisk = False
                    # '*'
                    elif ch == SIGN_ASTERISK:
                        is_asterisk = True

                # Inside of String Constant
                elif str_const != '':
                    if is_escape:
                        str_const += ch
                        is_escape = False
                    elif ch == SIGN_BACK_SLASH:
                        str_const += SIGN_BACK_SLASH
                        is_escape = True
                    else:
                        if is_single_q:
                            if ch == SIGN_SINGLE_QUOTATION:
                                str_const += SIGN_SINGLE_QUOTATION
                                tokens.append(str_const)
                                str_const = ''
                                is_ope = True
                                is_single_q = False
                            else:
                                str_const += ch
                        elif is_double_q:
                            if ch == SIGN_DOUBLE_QUOTATION:
                                str_const += SIGN_DOUBLE_QUOTATION
                                tokens.append(str_const)
                                str_const = ''
                                is_ope = True
                                is_double_q = False
                            else:
                                str_const += ch

                # '/'
                elif ch == SIGN_SLASH:
                    if str_token != '':
                        tokens.append(str_token)
                        str_token = ''
                        is_ope = True
                    if is_slash:
                        is_slash = False
                        break
                    is_slash = True

                # '*'
                elif ch == SIGN_ASTERISK:
                    if str_token != '':
                        tokens.append(str_token)
                        str_token = ''
                        is_ope = True
                    if is_slash:
                        is_comment = True
                        is_slash = False
                    else:
                        tokens.append(SIGN_ASTERISK)
                        is_ope = True

                # '\''
                elif ch == SIGN_SINGLE_QUOTATION:
                    if is_slash:
                        tokens.append(SIGN_SLASH)
                        is_ope = True
                        is_slash = False
                    str_const = SIGN_SINGLE_QUOTATION
                    is_single_q = True
                    is_escape = False

                # '"'
                elif ch == SIGN_DOUBLE_QUOTATION:
                    if is_slash:
                        tokens.append(SIGN_SLASH)
                        is_ope = True
                        is_slash = False
                    str_const = SIGN_DOUBLE_QUOTATION
                    is_double_q = True
                    is_escape = False

                # '\'
                elif ch == SIGN_BACK_SLASH:
                    if is_slash:
                        tokens.append(SIGN_SLASH)
                        is_ope = True
                        is_slash = False
                    elif is_double_q:
                        str_const = SIGN_DOUBLE_QUOTATION
                        continue
                    else:
                        if str_token != '':
                            tokens.append(str_token)
                            str_token = ''
                            is_ope = True
                    break

                # ' ', '\t', '　'
                elif ch in CH_DELIMITERS:
                    if is_slash:
                        tokens.append(SIGN_SLASH)
                        is_ope = True
                        is_slash = False
                    elif is_double_q:
                        str_const = SIGN_DOUBLE_QUOTATION + ch
                    else:
                        if str_token != '':
                            tokens.append(str_token)
                            str_token = ''
                            is_ope = True
                            is_double_q = False

                # Sign Marks
                elif ch in CH_SIGNS_OF_JAVA:
                    if is_slash:
                        tokens.append(SIGN_SLASH)
                        is_ope = True
                        is_slash = False
                    elif is_double_q:
                        str_const = SIGN_DOUBLE_QUOTATION + ch
                    else:
                        if str_token != '':
                            tokens.append(str_token)
                            str_token = ''
                        tokens.append(ch)
                        is_ope = True

                # Letters, Numbers
                else:
                    if is_slash:
                        tokens.append(SIGN_SLASH)
                        is_ope = True
                        is_slash = False
                    if is_double_q:
                        str_const = SIGN_DOUBLE_QUOTATION + ch
                    else:
                        str_token += ch

                pos_current += 1

            # End of One Line
            if is_slash:
                tokens.append(SIGN_SLASH)
                is_ope = True
            elif str_const != '':
                tokens.append(str_const)
                is_ope = True
            elif str_token != '':
                tokens.append(str_token)
                is_ope = True

            if fp is not None:
                strs = ''
                spc = ''
                for cnt, val in enumerate(tokens):
                    strs += spc + '[' + val + ']'
                    spc = ' '
                fp.write('%s %5d: %s\n' % ('|' if is_ope else ' ', num_lines, strs))

            if is_ope:
                num_steps += 1

        # End of All lines

    print('file encoding error in %s' % full_path_file, file=sys.stderr)
    return 0, 0, MSG_ERROR


# Scan SQL File
def scan_sql_file(full_path_file: str, fp) -> (int, int, str):

    if fp is not None:
        fp.write('%s\n' % full_path_file)

    for enc in ENCODINGS:

        num_lines = 0
        num_steps = 0
        is_comment = False

        file = open(full_path_file, 'r', encoding=enc)

        while True:

            try:
                str_line = file.readline()
            except Exception:       # noqa
                file.close()
                break

            # End of Data
            if not str_line:
                file.close()
                return num_lines, num_steps, MSG_NORMAL

            num_lines += 1

            str_line = str_line.rstrip('\n')  # for Display
            str_comp = str_line.strip()

            pos_current = 0
            pos_end = len(str_comp)

            tokens = []
            str_token = ''
            str_const = ''
            is_double_q = False
            num_single_q = 0
            is_minus = False
            is_slash = False
            is_asterisk = False
            is_ope = False

            while pos_current < pos_end:

                ch = str_comp[pos_current]

                # Inside of Comment
                if is_comment:
                    # before '*' was appeared
                    if is_asterisk:
                        # '*/'
                        if ch == SIGN_SLASH:
                            is_comment = False
                        # '**'
                        elif ch == SIGN_ASTERISK:
                            is_asterisk = True
                        # '*?"
                        else:
                            is_asterisk = False
                    # '*'
                    elif ch == SIGN_ASTERISK:
                        is_asterisk = True

                # Inside of String Constant
                elif str_const != '':
                    if num_single_q == 1:
                        if ch == SIGN_SINGLE_QUOTATION:
                            str_const += SIGN_SINGLE_QUOTATION
                            num_single_q = 2
                        else:
                            str_const += ch
                    elif num_single_q == 2:
                        if ch == SIGN_SINGLE_QUOTATION:
                            str_const += SIGN_SINGLE_QUOTATION
                            num_single_q = 1
                        else:
                            tokens.append(str_const)
                            str_const = ''
                            is_ope = True
                            num_single_q = 0
                            continue
                    elif is_double_q:
                        if ch == SIGN_DOUBLE_QUOTATION:
                            str_const += SIGN_DOUBLE_QUOTATION
                            tokens.append(str_const)
                            str_const = ''
                            is_ope = True
                            is_double_q = False
                        else:
                            str_const += ch

                # '/'
                elif ch == SIGN_SLASH:
                    if str_token != '':
                        tokens.append(str_token)
                        str_token = ''
                        is_ope = True
                    if is_slash:
                        is_slash = False
                        break
                    is_slash = True

                # '*'
                elif ch == SIGN_ASTERISK:
                    if str_token != '':
                        tokens.append(str_token)
                        str_token = ''
                        is_ope = True
                    if is_slash:
                        is_comment = True
                        is_slash = False
                    else:
                        tokens.append(SIGN_ASTERISK)
                        is_ope = True

                # '-'
                elif ch == SIGN_MINUS:
                    # Before '-' is appeared
                    if is_minus:
                        is_minus = False
                        break
                    is_minus = True

                # '\''
                elif ch == SIGN_SINGLE_QUOTATION:
                    if is_minus:
                        tokens.append(SIGN_MINUS)
                        is_minus = False
                    elif str_token != '':
                        tokens.append(str_token)
                        str_token = ''
                        is_ope = True
                    str_const = SIGN_SINGLE_QUOTATION
                    num_single_q = 1

                # '"'
                elif ch == SIGN_DOUBLE_QUOTATION:
                    if is_minus:
                        tokens.append(SIGN_MINUS)
                        is_minus = False
                    elif str_token != '':
                        tokens.append(str_token)
                        str_token = ''
                        is_ope = True
                    str_const = SIGN_DOUBLE_QUOTATION
                    is_double_q = True

                # ' ', '\t', '　'
                elif ch in CH_DELIMITERS:
                    if is_minus:
                        tokens.append(SIGN_MINUS)
                        is_minus = False
                    elif str_token != '':
                        tokens.append(str_token)
                        str_token = ''
                    is_ope = True

                # Sign Marks
                elif ch in CH_SIGNS_OF_JAVA:
                    if is_minus:
                        tokens.append(SIGN_MINUS)
                        is_minus = False
                    elif str_token != '':
                        tokens.append(str_token)
                        str_token = ''
                    tokens.append(ch)
                    is_ope = True

                # Letters, Numbers
                else:
                    str_token += ch

                pos_current += 1

            # End of One Line
            if is_slash:
                tokens.append(SIGN_SLASH)
                is_ope = True
            elif is_minus:
                tokens.append(SIGN_MINUS)
                is_ope = True
            elif str_const != '':
                tokens.append(str_const)
                is_ope = True
            elif str_token != '':
                tokens.append(str_token)
                is_ope = True

            if fp is not None:
                strs = ''
                spc = ''
                for cnt, val in enumerate(tokens):
                    strs += spc + '[' + val + ']'
                    spc = ' '
                fp.write('%s %5d: %s\n' % ('|' if is_ope else ' ', num_lines, strs))

            if is_ope:
                num_steps += 1

        # End of All lines

    print('file encoding error in %s' % full_path_file, file=sys.stderr)
    return 0, 0, MSG_ERROR


# Scan Text File
def scan_text_file(full_path_file: str) -> (int, int, str):

    for enc in ENCODINGS:

        num_lines = 0

        file = open(full_path_file, 'r', encoding=enc)

        while True:

            try:
                str_line = file.readline()
            except Exception:       # noqa
                file.close()
                break

            # End of Data
            if not str_line:
                file.close()
                return num_lines, None, MSG_NORMAL

            num_lines += 1

        # End of All lines

    print('file encoding error in %s' % full_path_file, file=sys.stderr)
    return None, None, MSG_ERROR


# Seek Directories
def seek_directories(write_excel: WriteExcel, level: int, dir_root: str, dir_relative: str, fp) -> None:

    dirs = []
    files = []

    for path in os.listdir(dir_root):

        if os.path.isfile(os.path.join(dir_root, path)):
            files.append(path)
        else:
            dirs.append(path)

    files.sort(key=str.lower)
    for file in files:
        base, ext = os.path.splitext(file)
        write_excel.write_cell(CELL_COL_NO, write_excel.get_row(), None, None, NUMBER_FORMAT)
        write_excel.write_cell(CELL_COL_PATH, dir_relative, ALIGN_LEFT_NO_WRAP, None, None)
        write_excel.write_cell(CELL_COL_FILE, file, ALIGN_LEFT_NO_WRAP, None, None)
        write_excel.write_cell(CELL_COL_EXT, ext, ALIGN_CENTER, None, None)
        # Ignore Files
        if (base.startswith('.') and ext == '') or ext in IGNORE_EXTENDS:
            lines = None
            steps = None
        elif ext == '.py':
            lines, steps, msg = scan_python_file(os.path.join(dir_root, file), fp)
        elif ext in ('.java', '.c', '.cpp'):
            lines, steps, msg = scan_java_file(os.path.join(dir_root, file), fp)
        elif ext == '.sql':
            lines, steps, msg = scan_sql_file(os.path.join(dir_root, file), fp)
        elif ext == '.txt':
            lines, steps, msg = scan_text_file(os.path.join(dir_root, file))
        # Other Files
        else:
            lines, steps, msg = scan_text_file(os.path.join(dir_root, file))
        write_excel.write_cell(CELL_COL_LINES, lines, None, None, NUMBER_FORMAT)
        write_excel.write_cell(CELL_COL_STEPS, steps, None, None, NUMBER_FORMAT)
        print('%s %s %s %s %s' % (dir_relative, file, ext,
                                  lines if lines is not None else '-', steps if steps is not None else '-'))
        write_excel.next_row()

    dirs.sort(key=str.lower)
    for dir_nest in dirs:
        seek_directories(write_excel, level + 1, os.path.join(dir_root, dir_nest), os.path.join(dir_relative, dir_nest), fp)

    return


# Get Current Time
def get_current_time() -> str:

    now = datetime.datetime.now()
    dt = now.strftime("%Y-%m-%d %H:%M:%S")
    return dt


# Main
def main() -> None:

    try:
        options, arguments = getopt.getopt(sys.argv[1:], shortopts="h", longopts=["help"])
    except getopt.error as message:
        print(message)
        print(__doc__)
        sys.exit(1)

    for option, argument in options:
        if option in ("-h", "--help"):
            print(__doc__)
            sys.exit(0)

    print('Source Code Counter - start [%s]' % get_current_time())

    # fp = None
    fp = open(OUT_DEBUG, 'w', encoding='utf-8')
    write_excel = WriteExcel(IN_EXCEL, OUT_EXCEL, OUT_SHEET)

    seek_directories(write_excel, 0, IN_SRC_ROOT + IN_SRC_RELATIVE, IN_SRC_RELATIVE, fp)

    write_excel.close()
    if fp is not None:
        fp.close()

    print('Source Code Counter - end [%s]' % get_current_time())

    sys.exit(0)


# Goto Main
if __name__ == '__main__':
    main()
